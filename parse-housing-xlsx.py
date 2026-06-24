#!/usr/bin/env python3
# LH 매입/전세임대 첨부 '주택목록(공급대상주택)' xlsx → 결정론 파서 (외부 LLM API 미사용)
#  - 헤더행 자동탐지(키워드 점수) + 2~3행 병합헤더 결합 + 컬럼 키워드 매핑
#  - 출력 1) data/derived/lh/<panId>/housing_list.json (호별 전체 + 컬럼매핑)
#         2) requirements.json 에 '주택목록' 요약 주입(매칭용: 지역분포/전용면적분포/임대조건/방수/주택형)
# 사용: python3 parse-housing-xlsx.py <panId> [<panId> ...]
#       python3 parse-housing-xlsx.py --all          (xlsx 보유 + requirements.json 있는 건 전체)
import sys, os, json, glob, re
import openpyxl

ROOT = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data')

# 컬럼 키워드 → 필드 (우선순위 순서로 매칭)
def classify(label):
    s = re.sub(r'\s+', '', label or '')
    if not s: return None
    # 면적 계열 (전용 우선, 공용/계 구분)
    if '전용' in s: return '전용면적'
    if '주거공용' in s or ('공용' in s and '면적' in s): return '주거공용면적'
    if ('면적' in s and ('계' in s or '합' in s)) : return '면적계'
    if '확장' in s and '면적' in s: return '확장면적'
    # 임대조건
    if '보증' in s: return '보증금'
    if '임대료' in s or ('월' in s and '세' in s) or '월세' in s: return '임대료'
    # 주택 식별
    if '도로명' in s or '소재지' in s or s == '주소' or '주소' in s: return '주소'
    if '지자체' in s or s == '지역' or '지역본부' in s: return '지자체'
    if '주택군' in s: return '주택군'
    if '공급형' in s or '주택형' in s or s == '형' or s.endswith('형'): return '주택형'
    if s == '동' or '동번호' in s: return '동'
    if s == '호' or '호수' in s: return '호'
    if '방' in s: return '방수'
    if '층' in s: return '층수'
    if '순번' in s or s == '번호' or s == 'no' or s == 'No'.lower(): return '순번'
    if '명' == s or s == '주택명': return '주택명'
    if '코드' in s: return '코드'
    return None

HEADER_KW = ['순번','지자체','주소','소재지','도로명','전용','면적','보증','임대료','임대조건',
             '공급형','주택형','주택군','방','층','동','호','지역본부']

def header_score(row):
    txt = ' '.join('' if c is None else str(c) for c in row)
    txt = re.sub(r'\s+', '', txt)
    return sum(1 for k in HEADER_KW if k in txt)

def to_num(v):
    if v is None: return None
    if isinstance(v,(int,float)): return float(v)
    s = re.sub(r'[^0-9.\-]', '', str(v))
    if s in ('', '.', '-'): return None
    try: return float(s)
    except: return None

def parse_sheet(ws):
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return None
    # 헤더행 탐지: 상단 25행 중 키워드 점수 최고(>=4)
    scan = min(25, len(rows))
    scores = [(header_score(rows[i]), i) for i in range(scan)]
    best_score, H = max(scores)
    if best_score < 4: return None
    # 데이터 시작행: H 이후, '순번/숫자'로 시작하거나 전용면적이 숫자인 첫 행
    width = max(len(r) for r in rows)
    def cell(r,c): return r[c] if c < len(r) else None
    # 임시 헤더 라벨(H행)으로 컬럼 후보 잡아 데이터 시작 탐지
    data_start = None
    for i in range(H+1, min(H+5, len(rows))):
        r = rows[i]
        first3 = [cell(r,c) for c in range(min(3,len(r)))]
        if any(isinstance(x,(int,float)) and float(x)==1 for x in first3):
            data_start = i; break
    if data_start is None:
        data_start = H+1
    # 헤더 라벨: H ~ data_start-1 행을 컬럼별로 결합(병합 2~3행 헤더 대응)
    labels = []
    for c in range(width):
        parts = []
        for i in range(H, data_start):
            v = cell(rows[i], c)
            if v is not None and str(v).strip():
                parts.append(str(v).strip())
        labels.append(' '.join(parts))
    # 컬럼 매핑(첫 매칭 우선, 중복 필드는 먼저 잡힌 컬럼 유지)
    colmap = {}
    for c,lab in enumerate(labels):
        f = classify(lab)
        if f and f not in colmap:
            colmap[f] = c
    # 데이터 읽기
    recs = []
    for i in range(data_start, len(rows)):
        r = rows[i]
        # 종료: 전용면적/순번 모두 비숫자이고 행 텍스트가 주석(*)이거나 거의 빈 행
        area = to_num(cell(r, colmap['전용면적'])) if '전용면적' in colmap else None
        seq  = cell(r, colmap['순번']) if '순번' in colmap else None
        nonempty = [x for x in r if x is not None and str(x).strip()]
        if area is None and not (isinstance(seq,(int,float))):
            # 데이터 영역 끝(주석/합계행)으로 판단 — 단 중간 빈행 1~2개는 건너뜀
            if len(nonempty) <= 1:
                continue
            txt = ''.join(str(x) for x in nonempty)
            if txt.startswith('*') or '합계' in txt or '소계' in txt or '계' == txt:
                continue
            # 면적 없는데 텍스트 많으면 비데이터 → 스킵
            continue
        rec = {}
        for f,c in colmap.items():
            v = cell(r,c)
            if f in ('전용면적','주거공용면적','면적계','확장면적','보증금','임대료','방수'):
                rec[f] = to_num(v)
            else:
                rec[f] = (str(v).strip() if v is not None else None)
        recs.append(rec)
    return {'header_row': H, 'data_start': data_start, 'colmap': colmap, 'labels': labels, 'records': recs}

SIDO = re.compile(r'(특별자치도|특별자치시|특별시|광역시|도|시)$')
def sigungu(addr, jicha):
    # 주소 2번째 토큰을 시군구로, 없으면 지자체명 괄호 안
    if addr:
        toks = str(addr).split()
        if len(toks) >= 2:
            # 토큰0이 시도, 토큰1이 시군구. 단 '세종'처럼 단일도 처리
            return toks[1]
        if toks: return toks[0]
    if jicha:
        m = re.search(r'[(（]([^)）]+)[)）]', str(jicha))
        if m: return m.group(1)
    return None

def summarize(panId, fname, sheetname, parsed):
    recs = parsed['records']
    areas = [r['전용면적'] for r in recs if r.get('전용면적')]
    deposits = [r['보증금'] for r in recs if r.get('보증금')]
    rents = [r['임대료'] for r in recs if r.get('임대료')]
    # 지역분포
    region = {}
    for r in recs:
        sg = sigungu(r.get('주소'), r.get('지자체') or r.get('주택명'))
        if sg: region[sg] = region.get(sg,0)+1
    # 전용면적 구간분포
    bands = {'~20㎡':0,'20~30㎡':0,'30~40㎡':0,'40~50㎡':0,'50㎡~':0}
    for a in areas:
        if a < 20: bands['~20㎡']+=1
        elif a < 30: bands['20~30㎡']+=1
        elif a < 40: bands['30~40㎡']+=1
        elif a < 50: bands['40~50㎡']+=1
        else: bands['50㎡~']+=1
    bands = {k:v for k,v in bands.items() if v}
    rooms = {}
    for r in recs:
        rm = r.get('방수')
        if rm: rooms[str(int(rm))+'개'] = rooms.get(str(int(rm))+'개',0)+1
    types = sorted({r['주택형'] for r in recs if r.get('주택형')})
    uniq_area = sorted({round(a,1) for a in areas})
    # 고유면적이 많으면(>40) 목록 생략 — 구간분포+최소/최대로 충분, 상세는 housing_list.json
    area_list = uniq_area if len(uniq_area) <= 40 else None
    return {
        '출처파일': fname,
        '시트': sheetname,
        '총호수': len(recs),
        '지역분포': dict(sorted(region.items(), key=lambda x:-x[1])),
        '전용면적': {
            '최소': round(min(areas),2) if areas else None,
            '최대': round(max(areas),2) if areas else None,
            '구간분포': bands,
            '고유면적수': len(uniq_area),
            '목록': area_list,
        },
        '임대조건': {
            '보증금': {'최소':int(min(deposits)),'최대':int(max(deposits))} if deposits else None,
            '월임대료': {'최소':int(min(rents)),'최대':int(max(rents))} if rents else None,
            '비고': None if (deposits or rents) else '첨부에 임대료 미기재(계약시 결정 등)',
        },
        '방수분포': dict(sorted(rooms.items())),
        '주택형': types,
    }

def find_xlsx(panId):
    fs = glob.glob(os.path.join(ROOT,'raw','lh',panId,'files','*.xlsx'))
    return fs[0] if fs else None

def process(panId):
    xlsx = find_xlsx(panId)
    if not xlsx: return (panId,'xlsx없음')
    fname = os.path.basename(xlsx)
    wb = openpyxl.load_workbook(xlsx, read_only=True, data_only=True)
    best = None
    for sn in wb.sheetnames:
        p = parse_sheet(wb[sn])
        if p and (best is None or len(p['records']) > len(best[1]['records'])):
            best = (sn, p)
    wb.close()
    if not best:
        return (panId,'헤더탐지실패')
    sheetname, parsed = best
    if not parsed['records']:
        return (panId,'데이터행0(매핑:'+','.join(parsed['colmap'].keys())+')')
    derived = os.path.join(ROOT,'derived','lh',panId)
    os.makedirs(derived, exist_ok=True)
    # housing_list.json
    hl = {'panId':panId,'출처파일':fname,'시트':sheetname,
          '컬럼매핑':parsed['colmap'],'헤더행':parsed['header_row'],
          '총호수':len(parsed['records']),'주택':parsed['records']}
    with open(os.path.join(derived,'housing_list.json'),'w',encoding='utf-8') as f:
        json.dump(hl,f,ensure_ascii=False,indent=2)
    # requirements.json 주입
    summary = summarize(panId, fname, sheetname, parsed)
    reqp = os.path.join(derived,'requirements.json')
    if os.path.exists(reqp):
        with open(reqp,encoding='utf-8') as f: req = json.load(f)
        req['주택목록'] = summary
        with open(reqp,'w',encoding='utf-8') as f:
            json.dump(req,f,ensure_ascii=False,indent=2)
    return (panId,f"OK {summary['총호수']}호 면적{summary['전용면적']['최소']}~{summary['전용면적']['최대']}㎡ "
                  f"임대료{'있음' if summary['임대조건']['월임대료'] else '없음'} 지역{len(summary['지역분포'])}곳")

def all_targets():
    out = []
    for d in sorted(glob.glob(os.path.join(ROOT,'derived','lh','*','requirements.json'))):
        panId = os.path.basename(os.path.dirname(d))
        if find_xlsx(panId): out.append(panId)
    return out

if __name__ == '__main__':
    args = sys.argv[1:]
    if not args:
        print('사용: parse-housing-xlsx.py <panId>... | --all'); sys.exit(1)
    targets = all_targets() if args[0]=='--all' else args
    okc=0
    for p in targets:
        pid,msg = process(p)
        flag = 'OK' if msg.startswith('OK') else '⚠️'
        if msg.startswith('OK'): okc+=1
        print(f"{flag} {pid}: {msg}")
    print(f"\n완료: {okc}/{len(targets)}")
